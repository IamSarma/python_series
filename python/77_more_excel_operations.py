import openpyxl
from openpyxl.styles import Font

# Setting the font style of cells
wb = openpyxl.Workbook()
sheet = wb["Sheet"]
# italic_24_font = Font(size=24, italic=True)
# sheet["A1"] = "Python is awesome!!!"
# sheet["A1"].font = italic_24_font
# wb.save("styles.xlsx")


# Font objecct(s)
# font_obj_1 = Font(name="Times New Roman", bold=True)
# sheet["A1"] = "Bold Times New Roman"
# sheet["A1"].font = font_obj_1

# font_obj_2 = Font(size=24, italic=True)
# sheet["B3"] = "24 pt Italic"
# sheet["B3"].font = font_obj_2

# wb.save("styles.xlsx")


# Formulas
# sheet["A1"] = 100
# sheet["A2"] = 200
# sheet["A3"] = "=SUM(A1:A2)"
# wb.save("write_formula.xlsx")


# Setting row height and column width
# sheet["A1"] = "Tall row"
# sheet["B2"] = "Wide column"
# sheet.row_dimensions[1].height = 70
# sheet.column_dimensions["B"].width = 20
# wb.save("row_column_dimensions.xlsx")


# Merging and Unmerging cells
sheet["A1"] = "Twelve cells merged together"
sheet["C5"] = "Two merged cells"

# Merging
sheet.merge_cells("A1:D3")
sheet.merge_cells("C5:D5")
wb.save("merged.xlsx")

# Unmerging
sheet.unmerge_cells("A1:D3")
sheet.unmerge_cells("C5:D5")
wb.save("un_merged.xlsx")
