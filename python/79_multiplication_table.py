# Program to generate multiplication table basis the user input
# and update the same in Excel
import openpyxl
import sys

# Create workbook
wb = openpyxl.Workbook()

# Create worksheet object
sheet = wb["Sheet"]

# User input
table_number = int(sys.argv[1]) + 2

# Generate table data and update in Excel
for i in range(2, table_number):
    sheet.cell(row=i, column=1).value = i - 1
    for j in range(2, table_number):
        sheet.cell(row=1, column=j).value = j - 1
        sheet.cell(row=i, column=j).value = sheet.cell(
            row=i, column=1).value * sheet.cell(row=1, column=j).value

# Save wokbook
wb.save("multiplication_table.xlsx")
