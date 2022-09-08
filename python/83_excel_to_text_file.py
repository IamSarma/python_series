# Program to copy data from spreadhsheet column(s) to individual text file(s)
# column A to one text file, column B to another and so on
import openpyxl
import os

# Folder path containing text file(s)
target_folder = os.getcwd() + "/text_files"

# Create workbook object
wb = openpyxl.load_workbook("text_to_excel.xlsx")

# Create sheet object
sheet = wb["Sheet1"]

# Get last/max row and last/max column to loop through data
last_row = sheet.max_row
last_column = sheet.max_column

# Loop through spreadsheet data, create and copy to text file(s)
for i in range(1, last_column + 1):
    text_file_name = str(i) + ".txt"
    text_file_obj = open(os.path.join(target_folder, text_file_name), "w")
    for j in range(1, last_row + 1):
        pass


# Save workbook
wb.save("text_to_excel.xlsx")
