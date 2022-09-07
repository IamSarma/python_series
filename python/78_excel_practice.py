# Excel practice question(s)
import openpyxl
from openpyxl.utils import column_index_from_string, get_column_letter


# Open/load workbook
wb = openpyxl.load_workbook("excel_practice.xlsx")

# Sheetnames
print(wb.sheetnames)

# Retrieve worksheet object
sheet = wb["Sheet1"]
print(sheet)

# Retrieve worksheet object for activesheet
act_sheet = wb.active
print(act_sheet)

# Set cell value
sheet["A1"] = "Python"

# Get cell value
print(sheet["A1"].value)

# Retrieve cell's row and column as integers
c = sheet["A1"]
print("Selected cell is A1")
print(f"Row number: {c.row}")
print(f"Column number: {c.column}")

# Get last row
print(f"Last/Max row: {sheet.max_row}")
print(f"The data type of sheet.max_row is: {type(sheet.max_row)}")

# Get last column
print(f"Last/Max column: {sheet.max_column}")
print(f"The data type of sheet.max_column is: {type(sheet.max_column)}")

# Get column index as integer value
col_int_index = column_index_from_string("C")
print(f"The index of column C is: {col_int_index}")

# Get column letter/string name from integer index
col_index = 8
col_str_letter = get_column_letter(col_index)
print(f"The string letter for column index {col_index} is: {col_str_letter}")

# Get a tuple of all the cell objects in specific range of cells
range_of_cell_values = sheet["A1:C1"]
print(range_of_cell_values)
print(type(range_of_cell_values))

# Set formula in a cell
sheet["G3"] = "=SUM(G1:G2)"

# Set row height
sheet.row_dimensions[5].height = 100

# Hide column
sheet.column_dimensions["C"].hidden = True

# Save workbook
wb.save("excel_practice.xlsx")
