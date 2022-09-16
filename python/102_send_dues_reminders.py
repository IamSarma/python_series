#! python 3
# send_dues_reminders.py
# Send payment reminder email(s) based on payment status in spreadsheet
import openpyxl
import smtplib
import sys


# open the spreadsheet and get the latest dues status
wb = openpyxl.load_workbook("dues_records.xlsx")
sheet = wb["Sheet1"]
last_col = sheet.max_column
latest_month = sheet.cell(row=1, column=last_col).value


# Check each member's payment status
unpaid_members = {}
for r in range(2, sheet.max_row + 1):
    payment_status = sheet.cell(row=r, column=last_col).value
    if payment_status != "paid":
        name = sheet.cell(row=r, column=1).value
        email_address = sheet.cell(row=r, column=2).value
        unpaid_members[name] = email_address


# Login to email account
user_email = sys.argv[1]
user_pwd = sys.argv[2]
smtp_obj = smtplib.SMTP("smtp.gmail.com", 587)
smtp_obj.ehlo()
smtp_obj.starttls()
smtp_obj.login(user_email, user_pwd)


# Send out reminder mail(s)
for name, email in unpaid_members.items():
    mail_subject = f"Subject: {latest_month} dues unpaid\n"
    mail_content = f"Dear {name},\nRecords show that you have not paid dues for {latest_month}, Please make this payment as soon as possible. Thank you!"
    mail_body = f"{mail_subject}{mail_content}"
    # Send mail and check send status
    print(f"Sending email to {email}")
    send_mail_status = smtp_obj.sendmail(user_email, email, mail_body)
    if send_mail_status != {}:
        print(f"Unable to send email to {email}: {send_mail_status}")

# Logout/Quit from SMTP
smtp_obj.quit()
