import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string

# Opening Excel file
wb = openpyxl.load_workbook("example.xlsx")
# print(type(wb))


# Getting sheets from the workbook
# print(wb.sheetnames)
# sheet = wb["Sheet3"]
# print(sheet)
# print(type(sheet))
# print(sheet.title)
# active_sheet = wb.active
# print(active_sheet)


# Getting cells from the sheets
sheet = wb["Sheet1"]
# print(sheet["A1"])
# print(sheet["A1"].value)


# Getting the row, column and value from the cell
# c = sheet["B1"]
# print(c.value)
# print(f"Row {c.row}, Column {c.column} is {c.value}")
# print(f"Cell {c.coordinate} is {c.value}")
# print(sheet["C1"].value)


# Using cell() method to point to specific cell
# print(sheet.cell(row=1, column=2))
# print(sheet.cell(row=1, column=2).value)


# Loop through every other row / odd row(s)
# for i in range(1, 8, 2):
#     print(i, sheet.cell(row=i, column=2).value)


# Getting the highest row and column number
# print(f"Number of rows containing data: {sheet.max_row}")
# print(f"Number of columns containing data: {sheet.max_column}")


# Converting between column letters and numbers
# print(get_column_letter(1))
# print(get_column_letter(2))
# print(get_column_letter(200))

# print(get_column_letter(sheet.max_column))
# print(column_index_from_string("A"))
# print(column_index_from_string("ZZ"))


# Getting rows and columns from the sheets
# print(tuple(sheet["A1":"C3"]))
# for row_of_cell_objects in sheet["A1":"C3"]:
#     for cell_object in row_of_cell_objects:
#         print(cell_object.coordinate, cell_object.value)
#     print("----- END OF ROW -----")

for cell_object in list(sheet.columns)[1]:
    print(cell_object.value)
