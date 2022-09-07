# Program to invert row and column of the cells in the spreadsheet
import openpyxl

cells_data = []
items_data = []

# Create workbook object
wb = openpyxl.load_workbook("invert_rows_columns.xlsx")

# Create worksheet object
sheet = wb["Sheet1"]
target_sheet = wb["Sheet2"]

# Read data from spreadsheet
for i in range(1, sheet.max_column+1):
    for j in range(1, sheet.max_row+1):
        items_data.append(sheet.cell(row=j, column=i).value)
    cells_data.append(list(items_data))
    items_data = []

# Write inverted data to spreadsheet
for i in range(0, len(cells_data)):
    for j in range(0, len(cells_data[i])):
        target_sheet.cell(row=i+1, column=j+1).value = cells_data[i][j]

# Save workbook
wb.save("invert_rows_columns.xlsx")
