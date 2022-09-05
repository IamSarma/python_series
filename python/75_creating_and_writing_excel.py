import openpyxl

# Creating Excel workbook
# wb = openpyxl.Workbook()
# print(wb.sheetnames)
# sheet = wb.active
# print(sheet.title)
# sheet.title = "master_data"
# print(sheet.title)
# print(wb.sheetnames)


# Saving Excel workbook
# wb = openpyxl.Workbook()
# sheet = wb.active
# sheet.title = "master_data"
# wb.save("created_with_python.xlsx")


# Creating and removing sheets
wb = openpyxl.Workbook()
print(wb.sheetnames)
wb.create_sheet()
print(wb.sheetnames)
# Create new sheet at index 0
wb.create_sheet(index=0, title="First Sheet")
print(wb.sheetnames)
wb.create_sheet(index=2, title="Middle Sheet")
print(wb.sheetnames)
